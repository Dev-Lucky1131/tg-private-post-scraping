import yt_dlp
import openpyxl
import os
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
def clean_url(url):
    if url.startswith("https://www.youtube.com/watch?v=https://www.youtube.com/watch?v="):
        return url.replace("https://www.youtube.com/watch?v=https://www.youtube.com/watch?v=", "https://www.youtube.com/watch?v=")
    return url
def get_channel_videos(channel_url):
    ydl_opts = {
        'quiet': True,
        'extract_flat': True,
        'skip_download': True,
    }
    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        info = ydl.extract_info(channel_url, download=False)
        if 'entries' not in info:
            print("Не вдалося знайти відео на каналі.")
            return []
        videos = []
        for entry in info['entries']:
            video_id = entry.get('id', '')
            if not video_id:
                continue
            videos.append((entry.get('title', 'Без назви'), f"https://www.youtube.com/watch?v={video_id}"))
        return list(reversed(videos))
def apply_conditional_formatting(sheet):
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_rule = FormulaRule(formula=["$D2=\"non terminé\""], stopIfTrue=True, fill=red_fill)
    green_rule = FormulaRule(formula=["$D2=\"terminé\""], stopIfTrue=True, fill=green_fill)
    sheet.conditional_formatting.add("D2:D1000", red_rule)
    sheet.conditional_formatting.add("D2:D1000", green_rule)
def update_excel_sheet(sheet, videos):
    # Get the last row with actual data
    last_row = 1  # Start from header row
    for row in range(sheet.max_row, 1, -1):
        if any(sheet.cell(row=row, column=i).value for i in range(1, 5)):
            last_row = row
            break

    # Get the last number from column A
    last_number = 0
    if last_row > 1:  # If there's data beyond the header
        last_number_cell = sheet.cell(row=last_row, column=1).value
        if isinstance(last_number_cell, str) and '=' in last_number_cell:
            # If it's a formula, evaluate it manually
            last_number = last_row - 1
        else:
            try:
                last_number = int(float(last_number_cell))
            except (ValueError, TypeError):
                last_number = last_row - 1

    # Create a set of existing video URLs
    existing_videos = {sheet.cell(row=i, column=3).value for i in range(2, last_row + 1) if sheet.cell(row=i, column=3).value}

    # Find videos that aren't in the sheet
    new_videos = [video for video in videos if video[1] not in existing_videos]
    
    if not new_videos:
        print("Нові відео не знайдено.")
        return

    # Add only new videos after the last row with data
    for idx, (title, url) in enumerate(new_videos, start=1):
        new_row = last_row + idx
        new_number = last_number + idx
        sheet.cell(row=new_row, column=1, value=new_number)  # Use direct number instead of formula
        sheet.cell(row=new_row, column=2, value=title)
        sheet.cell(row=new_row, column=3, value=url)
        sheet.cell(row=new_row, column=4, value="non terminé")  

    apply_conditional_formatting(sheet)
def process_channel(base_url, file_path, sheet_position, is_shorts):
    if not os.path.exists(file_path):
        # Create a new workbook if file doesn't exist
        workbook = openpyxl.Workbook()
        # Create two sheets: one for videos, one for shorts
        if len(workbook.sheetnames) == 1:
            workbook.create_sheet("Shorts")
            workbook.worksheets[0].title = "Videos"
        # Add headers to both sheets
        for sheet in workbook.worksheets:
            sheet['A1'] = "№"
            sheet['B1'] = "Title"
            sheet['C1'] = "URL"
            sheet['D1'] = "Status"
        workbook.save(file_path)
        print(f"Created new file: {file_path}")
    
    workbook = openpyxl.load_workbook(file_path)
    if sheet_position < 1 or sheet_position > len(workbook.sheetnames):
        print(f"Аркуш на позиції {sheet_position} не знайдено.")
        workbook.close()
        return
    sheet = workbook.worksheets[sheet_position - 1]
    videos_url = base_url.rstrip("/") + "/videos"
    shorts_url = base_url.rstrip("/") + "/shorts"
    channel_url = shorts_url if is_shorts else videos_url
    videos = get_channel_videos(channel_url)
    if not videos:
        print(f"Відео для аркуша на позиції {sheet_position} не знайдено.")
        workbook.close()
        return
    update_excel_sheet(sheet, videos)
    workbook.save(file_path)
    workbook.close()
    print(f"Аркуш на позиції {sheet_position} оновлено.")
if __name__ == "__main__":
    youtube_channel_url = "https://www.youtube.com/@TheperfectfrenchwithDylane"
    excel_file_path = "French_Dylane.xlsx"
    
    # Ensure the file exists before processing
    if not os.path.exists(excel_file_path):
        wb = openpyxl.Workbook()
        wb.save(excel_file_path)
    
    process_channel(youtube_channel_url, excel_file_path, 1, is_shorts=False)
    process_channel(youtube_channel_url, excel_file_path, 2, is_shorts=True)