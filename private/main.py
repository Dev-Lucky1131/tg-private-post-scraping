from telethon import TelegramClient
import asyncio
import os
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

api_id =   
api_hash = ""  
channel_identifier = ""  

script_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(script_dir, "повідомлення.xlsx")

client = TelegramClient("session_name", api_id, api_hash)

async def fetch_messages():
    """Fetch messages from both public and private Telegram channels."""
    await client.start()
    try:
        if channel_identifier.startswith("@"):  
            chat = await client.get_entity(channel_identifier)  
        else:
            chat = await client.get_entity(int(channel_identifier))  
        
        messages = await client.get_messages(chat, limit=None)
        messages = list(reversed(messages))
        filtered_messages = [msg for msg in messages if msg.text]

        await client.disconnect()
        return filtered_messages, chat
    except Exception as e:
        print(f"Error fetching messages: {e}")
        await client.disconnect()
        return [], None

def create_or_load_workbook(path):
    """Create or load an Excel workbook."""
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "повідомлення"
        
        headers = ["Номер", "Пост", "Посилання", "Статус"]
        ws.append(headers)

        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    return wb, ws

def update_excel(ws, messages, chat):
    """Update Excel sheet with new messages."""
    channel_username = getattr(chat, 'username', None)
    existing_links = set()

    for row in range(2, ws.max_row + 1):
        link = ws.cell(row=row, column=3).value
        if link:
            existing_links.add(link)

    new_entries = []
    for msg in messages:
        if channel_username and msg.id:
            link = f"https://t.me/{channel_username}/{msg.id}"
        elif chat.id and msg.id:
            link = f"https://t.me/c/{abs(chat.id)}/{msg.id}"  # For private channels
        else:
            link = ""
        if link not in existing_links:
            new_entries.append((msg.text, link))

    for (post_text, link) in new_entries:
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value="=ROW()-1")  # Auto-numbering
        cell_post = ws.cell(row=new_row, column=2, value=post_text)
        cell_post.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
        ws.cell(row=new_row, column=3, value=link)
        ws.cell(row=new_row, column=4, value=0)  # Ensure last row remains 0

    # Set column widths
    ws.column_dimensions['A'].width = 6    
    ws.column_dimensions['B'].width = 40   
    ws.column_dimensions['C'].width = 20   
    ws.column_dimensions['D'].width = 10   

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")

    # Apply borders to all cells
    thin_border = Border(left=Side(style='thin', color="000000"),
                         right=Side(style='thin', color="000000"),
                         top=Side(style='thin', color="000000"),
                         bottom=Side(style='thin', color="000000"))

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border

    # Apply conditional formatting only if there are enough rows
    if ws.max_row > 1:
        status_range = f"D2:D{ws.max_row}"
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        ws.conditional_formatting.add(status_range, CellIsRule(operator="equal", formula=["0"], fill=red_fill))

    return new_entries



async def process_messages():
    """Main function to fetch messages and update the Excel sheet."""
    messages, chat = await fetch_messages()
    if not chat:
        print("Failed to fetch messages. Check your access permissions.")
        return

    wb, ws = create_or_load_workbook(excel_path)
    new_entries = update_excel(ws, messages, chat)

    temp_excel_path = excel_path + ".temp"
    wb.save(temp_excel_path)
    wb.close()

    shutil.move(temp_excel_path, excel_path)

    print(f"Оновлено Excel-файл: {excel_path}")
    print(f"Додано {len(new_entries)} нових записів.")

if __name__ == "__main__":
    asyncio.run(process_messages())
